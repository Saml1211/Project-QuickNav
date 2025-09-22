"""
Realtime Processor - Immediate processing of document events

Features:
- Immediate document content extraction and analysis
- Incremental ML feature computation
- Version comparison and conflict resolution
- Change notification and alerting
- Integration with GUI for real-time updates
- Streaming updates to connected clients
"""

import asyncio
import logging
import time
from datetime import datetime
from typing import Dict, List, Optional, Any, Callable
import json
import hashlib
from pathlib import Path
from dataclasses import dataclass
from enum import Enum

from ...doc_navigator import DocumentParser, DocumentScanner, DocumentRanker
from ...database.database_manager import DatabaseManager

logger = logging.getLogger(__name__)


class ProcessingPriority(Enum):
    """Processing priority levels."""
    IMMEDIATE = 1   # User-initiated actions
    HIGH = 2       # Recent modifications
    NORMAL = 3     # Regular processing
    LOW = 4        # Background analysis


@dataclass
class ProcessingTask:
    """Represents a document processing task."""
    task_id: str
    file_path: str
    project_id: Optional[str]
    document_type: Optional[str]
    priority: ProcessingPriority
    event_type: str
    created_at: float
    retries: int = 0
    max_retries: int = 3
    error: Optional[str] = None

    def __lt__(self, other):
        """Priority queue comparison."""
        return self.priority.value < other.priority.value


class ContentExtractor:
    """Extracts content and metadata from documents."""

    def __init__(self, config: Dict[str, Any]):
        self.config = config
        self.max_content_length = config.get('max_content_length', 5000)
        self.extract_text = config.get('extract_text', True)
        self.extract_metadata = config.get('extract_metadata', True)

    async def extract_document_content(self, file_path: str) -> Dict[str, Any]:
        """Extract content and metadata from document."""
        try:
            content_data = {
                'file_path': file_path,
                'extracted_at': datetime.now(),
                'extraction_method': 'basic',
                'word_count': 0,
                'page_count': 0,
                'content_preview': '',
                'metadata': {},
                'text_content': '',
                'extraction_success': False
            }

            file_ext = Path(file_path).suffix.lower()

            if file_ext == '.pdf':
                content_data.update(await self._extract_pdf_content(file_path))
            elif file_ext in ['.docx', '.doc']:
                content_data.update(await self._extract_word_content(file_path))
            elif file_ext in ['.xlsx', '.xls']:
                content_data.update(await self._extract_excel_content(file_path))
            elif file_ext == '.txt':
                content_data.update(await self._extract_text_content(file_path))
            elif file_ext in ['.vsdx', '.vsd']:
                content_data.update(await self._extract_visio_metadata(file_path))
            else:
                # Basic file information only
                content_data.update(await self._extract_basic_info(file_path))

            return content_data

        except Exception as e:
            logger.error(f"Content extraction failed for {file_path}: {e}")
            return {
                'file_path': file_path,
                'extracted_at': datetime.now(),
                'extraction_success': False,
                'error': str(e)
            }

    async def _extract_pdf_content(self, file_path: str) -> Dict[str, Any]:
        """Extract content from PDF files."""
        try:
            # Try to import PDF processing libraries
            try:
                import PyPDF2
                pdf_available = True
            except ImportError:
                pdf_available = False

            if not pdf_available:
                return await self._extract_basic_info(file_path)

            with open(file_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)

                text_content = ""
                for page in reader.pages:
                    text_content += page.extract_text() + "\n"

                # Truncate if too long
                if len(text_content) > self.max_content_length:
                    content_preview = text_content[:self.max_content_length] + "..."
                else:
                    content_preview = text_content

                # Count words
                word_count = len(text_content.split())

                return {
                    'extraction_method': 'pypdf2',
                    'page_count': len(reader.pages),
                    'word_count': word_count,
                    'content_preview': content_preview,
                    'text_content': text_content if self.extract_text else '',
                    'metadata': dict(reader.metadata) if reader.metadata else {},
                    'extraction_success': True
                }

        except Exception as e:
            logger.debug(f"PDF extraction failed for {file_path}: {e}")
            return await self._extract_basic_info(file_path)

    async def _extract_word_content(self, file_path: str) -> Dict[str, Any]:
        """Extract content from Word documents."""
        try:
            # Try to import Word processing libraries
            try:
                import docx
                docx_available = True
            except ImportError:
                docx_available = False

            if not docx_available or not file_path.endswith('.docx'):
                return await self._extract_basic_info(file_path)

            doc = docx.Document(file_path)

            # Extract text content
            text_content = ""
            for paragraph in doc.paragraphs:
                text_content += paragraph.text + "\n"

            # Truncate if too long
            if len(text_content) > self.max_content_length:
                content_preview = text_content[:self.max_content_length] + "..."
            else:
                content_preview = text_content

            # Count words
            word_count = len(text_content.split())

            # Extract metadata
            metadata = {}
            if doc.core_properties:
                props = doc.core_properties
                metadata.update({
                    'author': props.author,
                    'title': props.title,
                    'subject': props.subject,
                    'created': str(props.created) if props.created else None,
                    'modified': str(props.modified) if props.modified else None,
                    'last_modified_by': props.last_modified_by
                })

            return {
                'extraction_method': 'python-docx',
                'word_count': word_count,
                'content_preview': content_preview,
                'text_content': text_content if self.extract_text else '',
                'metadata': metadata,
                'extraction_success': True
            }

        except Exception as e:
            logger.debug(f"Word extraction failed for {file_path}: {e}")
            return await self._extract_basic_info(file_path)

    async def _extract_excel_content(self, file_path: str) -> Dict[str, Any]:
        """Extract content from Excel files."""
        try:
            # Try to import Excel processing libraries
            try:
                import openpyxl
                excel_available = True
            except ImportError:
                excel_available = False

            if not excel_available or not file_path.endswith('.xlsx'):
                return await self._extract_basic_info(file_path)

            workbook = openpyxl.load_workbook(file_path, data_only=True)

            # Extract sheet names and basic info
            sheet_names = workbook.sheetnames
            total_cells = 0
            content_preview = ""

            for sheet_name in sheet_names[:3]:  # Limit to first 3 sheets
                sheet = workbook[sheet_name]
                sheet_content = f"Sheet: {sheet_name}\n"

                # Sample first few rows
                for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
                    if any(cell is not None for cell in row):
                        row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                        sheet_content += row_text + "\n"
                        total_cells += len([c for c in row if c is not None])

                content_preview += sheet_content + "\n"

            # Truncate if too long
            if len(content_preview) > self.max_content_length:
                content_preview = content_preview[:self.max_content_length] + "..."

            return {
                'extraction_method': 'openpyxl',
                'sheet_count': len(sheet_names),
                'sheet_names': sheet_names,
                'cell_count': total_cells,
                'content_preview': content_preview,
                'metadata': {
                    'sheet_names': sheet_names
                },
                'extraction_success': True
            }

        except Exception as e:
            logger.debug(f"Excel extraction failed for {file_path}: {e}")
            return await self._extract_basic_info(file_path)

    async def _extract_text_content(self, file_path: str) -> Dict[str, Any]:
        """Extract content from text files."""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                text_content = file.read()

            # Truncate if too long
            if len(text_content) > self.max_content_length:
                content_preview = text_content[:self.max_content_length] + "..."
            else:
                content_preview = text_content

            word_count = len(text_content.split())

            return {
                'extraction_method': 'text',
                'word_count': word_count,
                'content_preview': content_preview,
                'text_content': text_content if self.extract_text else '',
                'extraction_success': True
            }

        except Exception as e:
            logger.debug(f"Text extraction failed for {file_path}: {e}")
            return await self._extract_basic_info(file_path)

    async def _extract_visio_metadata(self, file_path: str) -> Dict[str, Any]:
        """Extract metadata from Visio files."""
        # Visio files are complex binary/XML formats
        # For now, just return basic file info
        return await self._extract_basic_info(file_path)

    async def _extract_basic_info(self, file_path: str) -> Dict[str, Any]:
        """Extract basic file information."""
        try:
            import os
            stat = os.stat(file_path)

            return {
                'extraction_method': 'basic',
                'file_size': stat.st_size,
                'modified_time': datetime.fromtimestamp(stat.st_mtime),
                'extraction_success': True
            }

        except Exception as e:
            return {
                'extraction_method': 'basic',
                'extraction_success': False,
                'error': str(e)
            }


class VersionAnalyzer:
    """Analyzes document versions and detects changes."""

    def __init__(self, db_manager: DatabaseManager):
        self.db_manager = db_manager

    async def analyze_version_changes(self, file_path: str, project_id: str,
                                    document_type: str) -> Dict[str, Any]:
        """Analyze version changes for a document."""
        try:
            # Get existing versions from database
            existing_versions = await self.db_manager.get_document_versions(
                project_id, document_type
            )

            current_filename = Path(file_path).name
            parser = DocumentParser()
            current_metadata = parser.parse_filename(current_filename)

            analysis = {
                'is_new_version': False,
                'is_newer_than_existing': False,
                'version_increment': None,
                'status_change': None,
                'conflicts': [],
                'recommendations': []
            }

            if not existing_versions:
                analysis['is_new_version'] = True
                analysis['recommendations'].append("First version of this document type")
                return analysis

            # Compare with existing versions
            current_version = current_metadata.get('version')
            current_version_type = current_metadata.get('version_type')

            for existing in existing_versions:
                existing_version = existing.get('version_numeric')
                existing_filename = existing.get('filename')

                # Check for exact filename match (potential overwrite)
                if existing_filename == current_filename:
                    analysis['conflicts'].append(f"Filename conflict with existing: {existing_filename}")

                # Version comparison
                if current_version and existing_version:
                    if current_version > existing_version:
                        analysis['is_newer_than_existing'] = True
                        analysis['version_increment'] = current_version - existing_version

            # Status analysis
            current_status = current_metadata.get('status_tags', set())
            if current_status:
                # Check if this introduces a new status
                existing_statuses = set()
                for existing in existing_versions:
                    existing_tags = existing.get('status_tags', [])
                    if isinstance(existing_tags, str):
                        existing_tags = json.loads(existing_tags)
                    existing_statuses.update(existing_tags)

                new_statuses = current_status - existing_statuses
                if new_statuses:
                    analysis['status_change'] = list(new_statuses)

            # Generate recommendations
            if analysis['is_newer_than_existing']:
                analysis['recommendations'].append("Newer version detected - consider archiving older versions")

            if analysis['conflicts']:
                analysis['recommendations'].append("Resolve filename conflicts before proceeding")

            if 'AS-BUILT' in current_status:
                analysis['recommendations'].append("As-built document detected - high priority for indexing")

            return analysis

        except Exception as e:
            logger.error(f"Version analysis failed for {file_path}: {e}")
            return {'error': str(e)}


class RealtimeProcessor:
    """
    Real-time document processing system.

    Handles immediate processing of document events including:
    - Content extraction and analysis
    - Version comparison and conflict detection
    - ML feature computation
    - Change notifications
    """

    def __init__(self, config: Dict[str, Any], db_manager: DatabaseManager):
        self.config = config
        self.db_manager = db_manager

        # Configuration
        self.processing_timeout = config.get('processing_timeout', 30.0)
        self.max_concurrent_tasks = config.get('max_concurrent_tasks', 5)
        self.enable_content_extraction = config.get('enable_content_extraction', True)
        self.enable_ml_features = config.get('enable_ml_features', False)

        # Components
        self.content_extractor = ContentExtractor(config.get('content_extraction', {}))
        self.version_analyzer = VersionAnalyzer(db_manager)

        # Task management
        self.task_queue: asyncio.PriorityQueue = asyncio.PriorityQueue()
        self.active_tasks: Dict[str, asyncio.Task] = {}
        self.semaphore = asyncio.Semaphore(self.max_concurrent_tasks)

        # State management
        self.running = False
        self.processor_task: Optional[asyncio.Task] = None

        # Callbacks and notifications
        self.change_callbacks: List[Callable] = []
        self.update_callbacks: List[Callable] = []

        # Performance metrics
        self.metrics = {
            'tasks_processed': 0,
            'tasks_failed': 0,
            'content_extracted': 0,
            'versions_analyzed': 0,
            'avg_processing_time_ms': 0.0,
            'last_activity': time.time()
        }

    async def start(self):
        """Start the realtime processor."""
        if self.running:
            logger.warning("Realtime processor already running")
            return

        logger.info("Starting realtime processor...")

        self.processor_task = asyncio.create_task(self._processor_loop())
        self.running = True

        logger.info("Realtime processor started")

    async def shutdown(self):
        """Shutdown the realtime processor."""
        if not self.running:
            return

        logger.info("Shutting down realtime processor...")
        self.running = False

        # Cancel processor task
        if self.processor_task:
            self.processor_task.cancel()

        # Cancel active tasks
        if self.active_tasks:
            for task in self.active_tasks.values():
                task.cancel()

            # Wait for tasks to complete
            try:
                await asyncio.wait_for(
                    asyncio.gather(*self.active_tasks.values(), return_exceptions=True),
                    timeout=10
                )
            except asyncio.TimeoutError:
                logger.warning("Some tasks did not complete within timeout")

        logger.info("Realtime processor shutdown complete")

    async def process_event(self, event_type: str, file_path: str,
                          project_id: Optional[str] = None,
                          priority: ProcessingPriority = ProcessingPriority.NORMAL):
        """Submit a document event for processing."""
        if not self.running:
            logger.warning("Processor not running, ignoring event")
            return

        task_id = f"{event_type}_{hashlib.md5(file_path.encode()).hexdigest()[:8]}_{int(time.time())}"

        # Extract document type from file path
        from ...doc_navigator import DocumentTypeClassifier
        classifier = DocumentTypeClassifier()
        type_config = classifier.classify_document(file_path)
        document_type = type_config['type'] if type_config else None

        task = ProcessingTask(
            task_id=task_id,
            file_path=file_path,
            project_id=project_id,
            document_type=document_type,
            priority=priority,
            event_type=event_type,
            created_at=time.time()
        )

        await self.task_queue.put((priority.value, time.time(), task))
        logger.debug(f"Queued processing task: {task_id}")

    async def _processor_loop(self):
        """Main processing loop."""
        logger.info("Processor loop started")

        while self.running:
            try:
                # Get next task
                try:
                    priority, queued_time, task = await asyncio.wait_for(
                        self.task_queue.get(),
                        timeout=1.0
                    )
                except asyncio.TimeoutError:
                    continue

                # Process task asynchronously
                processing_task = asyncio.create_task(
                    self._process_task_with_semaphore(task)
                )
                self.active_tasks[task.task_id] = processing_task

                # Clean up completed tasks
                completed_tasks = [
                    task_id for task_id, t in self.active_tasks.items()
                    if t.done()
                ]
                for task_id in completed_tasks:
                    del self.active_tasks[task_id]

            except asyncio.CancelledError:
                break
            except Exception as e:
                logger.error(f"Processor loop error: {e}")
                await asyncio.sleep(1)

    async def _process_task_with_semaphore(self, task: ProcessingTask):
        """Process a task with semaphore-controlled concurrency."""
        async with self.semaphore:
            await self._process_task(task)

    async def _process_task(self, task: ProcessingTask):
        """Process a single document task."""
        start_time = time.time()

        try:
            logger.debug(f"Processing task {task.task_id}: {task.event_type} {task.file_path}")

            if task.event_type == 'deleted':
                await self._handle_deletion(task)
            else:
                await self._handle_update(task)

            # Update metrics
            processing_time_ms = (time.time() - start_time) * 1000
            self.metrics['tasks_processed'] += 1
            self.metrics['avg_processing_time_ms'] = (
                (self.metrics['avg_processing_time_ms'] * (self.metrics['tasks_processed'] - 1) +
                 processing_time_ms) / self.metrics['tasks_processed']
            )
            self.metrics['last_activity'] = time.time()

            logger.debug(f"Completed task {task.task_id} in {processing_time_ms:.1f}ms")

        except Exception as e:
            task.error = str(e)
            self.metrics['tasks_failed'] += 1
            logger.error(f"Task {task.task_id} failed: {e}")

            # Retry logic
            if task.retries < task.max_retries:
                task.retries += 1
                await asyncio.sleep(min(2 ** task.retries, 60))  # Exponential backoff
                await self.task_queue.put((task.priority.value, time.time(), task))
                logger.info(f"Retrying task {task.task_id} (attempt {task.retries})")

    async def _handle_update(self, task: ProcessingTask):
        """Handle document creation or modification."""
        if not os.path.exists(task.file_path):
            logger.warning(f"File no longer exists: {task.file_path}")
            return

        results = {
            'task_id': task.task_id,
            'file_path': task.file_path,
            'project_id': task.project_id,
            'document_type': task.document_type,
            'processed_at': datetime.now(),
            'content_extraction': None,
            'version_analysis': None,
            'ml_features': None
        }

        # Content extraction
        if self.enable_content_extraction:
            try:
                content_data = await self.content_extractor.extract_document_content(task.file_path)
                results['content_extraction'] = content_data

                if content_data.get('extraction_success'):
                    self.metrics['content_extracted'] += 1

                    # Update document in database with extracted content
                    await self._update_document_content(task, content_data)

            except Exception as e:
                logger.error(f"Content extraction failed for {task.file_path}: {e}")
                results['content_extraction'] = {'error': str(e)}

        # Version analysis
        if task.project_id and task.document_type:
            try:
                version_analysis = await self.version_analyzer.analyze_version_changes(
                    task.file_path, task.project_id, task.document_type
                )
                results['version_analysis'] = version_analysis
                self.metrics['versions_analyzed'] += 1

                # Handle version conflicts
                if version_analysis.get('conflicts'):
                    await self._handle_version_conflicts(task, version_analysis)

            except Exception as e:
                logger.error(f"Version analysis failed for {task.file_path}: {e}")
                results['version_analysis'] = {'error': str(e)}

        # ML feature computation
        if self.enable_ml_features and results['content_extraction']:
            try:
                ml_features = await self._compute_ml_features(task, results['content_extraction'])
                results['ml_features'] = ml_features
            except Exception as e:
                logger.error(f"ML feature computation failed for {task.file_path}: {e}")
                results['ml_features'] = {'error': str(e)}

        # Notify callbacks
        await self._notify_callbacks(task, results)

    async def _handle_deletion(self, task: ProcessingTask):
        """Handle document deletion."""
        # Mark as inactive in database
        await self.db_manager.execute_write(
            "UPDATE documents SET is_active = 0, updated_at = CURRENT_TIMESTAMP WHERE file_path = ?",
            [task.file_path]
        )

        # Notify callbacks
        await self._notify_callbacks(task, {
            'task_id': task.task_id,
            'file_path': task.file_path,
            'event_type': 'deleted',
            'processed_at': datetime.now()
        })

    async def _update_document_content(self, task: ProcessingTask, content_data: Dict[str, Any]):
        """Update document record with extracted content."""
        update_data = {}

        if content_data.get('word_count'):
            update_data['word_count'] = content_data['word_count']

        if content_data.get('page_count'):
            update_data['page_count'] = content_data['page_count']

        if content_data.get('content_preview'):
            update_data['content_preview'] = content_data['content_preview']

        if update_data:
            # Build dynamic update query
            set_clauses = [f"{key} = ?" for key in update_data.keys()]
            query = f"UPDATE documents SET {', '.join(set_clauses)}, updated_at = CURRENT_TIMESTAMP WHERE file_path = ?"
            params = list(update_data.values()) + [task.file_path]

            await self.db_manager.execute_write(query, params)

    async def _handle_version_conflicts(self, task: ProcessingTask, version_analysis: Dict[str, Any]):
        """Handle version conflicts detected during analysis."""
        conflicts = version_analysis.get('conflicts', [])

        for conflict in conflicts:
            logger.warning(f"Version conflict detected for {task.file_path}: {conflict}")

            # Could implement automatic conflict resolution here
            # For now, just log and notify

        # Notify about conflicts
        for callback in self.change_callbacks:
            try:
                if asyncio.iscoroutinefunction(callback):
                    await callback({
                        'type': 'version_conflict',
                        'file_path': task.file_path,
                        'conflicts': conflicts,
                        'analysis': version_analysis
                    })
                else:
                    callback({
                        'type': 'version_conflict',
                        'file_path': task.file_path,
                        'conflicts': conflicts,
                        'analysis': version_analysis
                    })
            except Exception as e:
                logger.warning(f"Conflict callback failed: {e}")

    async def _compute_ml_features(self, task: ProcessingTask, content_data: Dict[str, Any]) -> Dict[str, Any]:
        """Compute ML features from extracted content."""
        # Placeholder for ML feature computation
        # This would integrate with the ML pipeline when available

        features = {
            'document_length_category': self._categorize_document_length(content_data.get('word_count', 0)),
            'has_technical_content': self._detect_technical_content(content_data.get('content_preview', '')),
            'modification_frequency': await self._calculate_modification_frequency(task.file_path),
            'computed_at': datetime.now()
        }

        return features

    def _categorize_document_length(self, word_count: int) -> str:
        """Categorize document by length."""
        if word_count < 100:
            return 'very_short'
        elif word_count < 500:
            return 'short'
        elif word_count < 2000:
            return 'medium'
        elif word_count < 5000:
            return 'long'
        else:
            return 'very_long'

    def _detect_technical_content(self, content: str) -> bool:
        """Simple heuristic to detect technical content."""
        technical_keywords = [
            'specification', 'technical', 'diagram', 'schematic',
            'requirement', 'protocol', 'interface', 'configuration'
        ]

        content_lower = content.lower()
        return any(keyword in content_lower for keyword in technical_keywords)

    async def _calculate_modification_frequency(self, file_path: str) -> float:
        """Calculate how frequently this document is modified."""
        try:
            # Query modification history from database
            results = await self.db_manager.execute_query(
                """
                SELECT COUNT(*) as mod_count,
                       MIN(updated_at) as first_seen,
                       MAX(updated_at) as last_seen
                FROM documents
                WHERE file_path = ?
                """,
                [file_path]
            )

            if results and results[0]['mod_count'] > 1:
                first_seen = datetime.fromisoformat(results[0]['first_seen'])
                last_seen = datetime.fromisoformat(results[0]['last_seen'])
                days_span = (last_seen - first_seen).days

                if days_span > 0:
                    return results[0]['mod_count'] / days_span

            return 0.0

        except Exception as e:
            logger.debug(f"Error calculating modification frequency: {e}")
            return 0.0

    async def _notify_callbacks(self, task: ProcessingTask, results: Dict[str, Any]):
        """Notify registered callbacks of processing results."""
        for callback in self.update_callbacks:
            try:
                if asyncio.iscoroutinefunction(callback):
                    await callback(results)
                else:
                    callback(results)
            except Exception as e:
                logger.warning(f"Update callback failed: {e}")

    # Public API methods

    def register_change_callback(self, callback: Callable):
        """Register callback for document changes."""
        self.change_callbacks.append(callback)

    def register_update_callback(self, callback: Callable):
        """Register callback for processing updates."""
        self.update_callbacks.append(callback)

    async def get_status(self) -> Dict[str, Any]:
        """Get processor status."""
        return {
            "running": self.running,
            "queue_size": self.task_queue.qsize(),
            "active_tasks": len(self.active_tasks),
            "metrics": self.metrics.copy(),
            "config": {
                "content_extraction_enabled": self.enable_content_extraction,
                "ml_features_enabled": self.enable_ml_features,
                "max_concurrent_tasks": self.max_concurrent_tasks
            }
        }

    async def get_metrics(self) -> Dict[str, Any]:
        """Get performance metrics."""
        return self.metrics.copy()

    async def health_check(self) -> Dict[str, Any]:
        """Perform health check."""
        healthy = True
        issues = []

        if not self.running:
            healthy = False
            issues.append("Processor not running")

        # Check for stalled tasks
        if self.active_tasks:
            current_time = time.time()
            stalled_tasks = [
                task_id for task_id, task in self.active_tasks.items()
                if (current_time - task._callbacks[0].__self__.created_at) > self.processing_timeout
            ]

            if stalled_tasks:
                issues.append(f"Stalled tasks detected: {len(stalled_tasks)}")

        # Check error rate
        total_tasks = self.metrics['tasks_processed'] + self.metrics['tasks_failed']
        if total_tasks > 0:
            error_rate = self.metrics['tasks_failed'] / total_tasks
            if error_rate > 0.1:  # 10% error rate
                healthy = False
                issues.append(f"High error rate: {error_rate:.1%}")

        return {
            "healthy": healthy,
            "issues": issues,
            "metrics": self.metrics.copy()
        }